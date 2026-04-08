"""
TCI Menu Craft — Excel Grid Parser
Reads TCI_Banquet_Grid.xlsx and returns a structured dict of menu types,
pricing tiers, sections, subcategories with max-item limits, and notes.
"""

import openpyxl
from config import (
    EXCEL_PATH, EXCEL_SHEET, MENU_TYPE_ROW, PRICE_ROWS, MG_ROW,
    DATA_START_ROW, DATA_END_ROW, NOTES_START_ROW, MENU_TYPE_START_COL,
)


def parse_grid(filepath=None):
    """Parse the Banquet Grid Excel and return structured data.

    Returns:
        {
          "menu_types": {
            "CORPORATE (Veg)": {
              "pricing": {
                "Desire": {"price": 1200, "mg": 100},
                "Wish":   {"price": 1000, "mg": 100},
                "Walk":   {"price": 800,  "mg": 100},
              },
              "sections": {
                "Break-up": {
                  "Welcome Drink": {"max": 0},
                  ...
                },
                ...
              }
            },
            ...
          },
          "notes": ["Note 1...", ...]
        }
    """
    filepath = filepath or EXCEL_PATH
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[EXCEL_SHEET]

    # ── Read menu type names from row 6, columns B onward ────────────────
    menu_names = []
    col = MENU_TYPE_START_COL
    while True:
        val = ws.cell(row=MENU_TYPE_ROW, column=col).value
        if val is None:
            break
        menu_names.append((col, str(val).strip()))
        col += 1

    # ── Build pricing for each menu type ─────────────────────────────────
    menu_types = {}
    for col_idx, name in menu_names:
        mg_val = ws.cell(row=MG_ROW, column=col_idx).value or 0
        pricing = {}
        for tier, row_num in PRICE_ROWS.items():
            price = ws.cell(row=row_num, column=col_idx).value or 0
            pricing[tier] = {"price": int(price), "mg": int(mg_val)}
        menu_types[name] = {"pricing": pricing, "sections": {}}

    # ── Parse sections and subcategories (rows 13–35) ────────────────────
    current_section = None
    for row_num in range(DATA_START_ROW, DATA_END_ROW + 1):
        label = ws.cell(row=row_num, column=1).value
        if label is None:
            continue
        label = str(label).strip()
        if not label:
            continue

        # Check if this is a section header (no numeric data in col B+)
        has_data = False
        for col_idx, _ in menu_names:
            val = ws.cell(row=row_num, column=col_idx).value
            if val is not None:
                has_data = True
                break

        if not has_data:
            current_section = label
            for name in menu_types:
                menu_types[name]["sections"][current_section] = {}
        else:
            if current_section is None:
                continue
            for col_idx, mt_name in menu_names:
                val = ws.cell(row=row_num, column=col_idx).value
                max_items = int(val) if val is not None else 0
                menu_types[mt_name]["sections"][current_section][label] = {
                    "max": max_items
                }

    # ── Extract notes ────────────────────────────────────────────────────
    notes = []
    row = NOTES_START_ROW
    while row <= ws.max_row:
        val = ws.cell(row=row, column=1).value
        if val is not None:
            text = str(val).strip()
            if text and text.lower() != "notes:":
                notes.append(text)
        row += 1

    wb.close()
    return {"menu_types": menu_types, "notes": notes}


def run_diagnostic(filepath=None):
    """Print diagnostic info about the grid for debugging."""
    filepath = filepath or EXCEL_PATH
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[EXCEL_SHEET]

    print(f"Sheet: {EXCEL_SHEET}")
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    print("\n--- First 45 rows ---")
    for row in ws.iter_rows(min_row=1, max_row=min(45, ws.max_row), values_only=False):
        vals = [(c.column, c.value) for c in row if c.value is not None]
        if vals:
            print(f"  Row {row[0].row}: {vals}")

    print("\n--- Parsed Data ---")
    data = parse_grid(filepath)
    for mt_name, mt_data in data["menu_types"].items():
        print(f"\n{mt_name}:")
        for tier, info in mt_data["pricing"].items():
            print(f"  {tier}: ₹{info['price']}/head, MG={info['mg']}")
        for section, cats in mt_data["sections"].items():
            print(f"  [{section}]")
            for cat, info in cats.items():
                print(f"    {cat}: max={info['max']}")

    if data["notes"]:
        print("\nNotes:")
        for n in data["notes"]:
            print(f"  {n}")

    wb.close()


if __name__ == "__main__":
    run_diagnostic()
